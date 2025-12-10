# coidng=utf-8


from openpyxl import load_workbook


# res=sheet.cell(1,1)
# 错误演示：只演示到单元格，并未取值，cell是单元格

# 怎么拿到第一行的所有数据
class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]  # 打开一个表单名的表单，拿到对象后存到sheet_obj中
        self.max_row=self.sheet_obj.max_row
        # 获取一个表单对象

    def get_data(self, i, j):
        """根据传入的坐标来获取值"""
        return self.sheet_obj.cell(i, j).value


if __name__ == "__main__":
    res = DoExcel("xh.xlsx", "Python").get_data(1, 1)
    print(res)
