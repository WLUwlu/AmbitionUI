#coding=utf-8
from openpyxl import load_workbook

wb=load_workbook('xh.xlsx')
sheet=wb["Python"]

#res=sheet.cell(1,1)
#错误演示：只演示到单元格，并未取值，cell是单元格

#怎么拿到第一行的所有数据
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    #方法3：表格xlsx增加titile
    def get_header(self):
        """获取第一行的标题行"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]


        header=self.get_header()#拿到header
        test_data = []

        header=[] #存储我们的标题行
        for j in range(1,sheet.max_column+1)
            header.append(1,j)

        return header

    def get_data(self):
        '''根据嵌套循环读取数据'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name] #错误点二，不要写成字符串
        test_data=[]
        for i in range(1,sheet.max_row+1):  #(1,5) 1234
            sub_data={}
            #第一种
            sub_data["method"]=sheet.cell(i,1).value
            sub_data["url"]=sheet.cell(i,2).value
            sub_data["data"]=sheet.cell(i,3).value
            sub_data["expected"]=sheet.cell(i,4).value
            #第二种
            for j in range(2,sheet.max_column+1):
                sub_data[header[j]]=sheet.cell(i,j).value
            test_data.append(sub_data)
        return test_data #返回获取到的数据

if __name__=="__main__":
    print(DoExcel("xh.xlsx","Python").get_data())


