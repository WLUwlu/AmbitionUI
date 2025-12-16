#coding=utf-8
from openpyxl import load_workbook
from do_ddt_learn.read_config import ToolsConfig

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name


    def get_data(self,button="all"):
        """button:控制是否执行所有的用例，默认值为all，等于all的就会执行全部用例，不等于all就会进入分支判断
        button的值，只能输入all和列表这俩种类型的参数
        """
        #从配置文件读取mode值
        mode=ToolsConfig().read_config("tools.config","MODE","mode")
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]
        for i in range(2,sheet.max_row+1)
            sub_data={}
            sub_data["case_id"]=sheet.cell(i,2).value
            sub_data["module"]=sheet.cell(i,2).value
            sub_data["title"]=sheet.cell(i,2).value
            sub_data["method"]=sheet.cell(i,2).value
            sub_data["url"]=sheet.cell(i,2).value
            sub_data["data"]=sheet.cell(i,2).value
            sub_data["expected"]=sheet.cell(i,2).value
            test_data.append(sub_data)  #存储了所有的数据

        #根据button值进行判断
        if mode=="all":  #执行所有的用例
            final_data=test_data
        else:
            final_data=[]
            for item in test_data:  #对test_data所有的数据进行遍历
                if item["case_id"] in eval(mode):  #对mode进行转字符
                    final_data.append(item)

        return final_data #返回获取到的数据

        return test_data    #返回获取到的数据

    if __name__=="__main__":
        print(DoExcel("xh.xlsx","python").get_data())



