#存到excel里面python去操作excel
#xlsx----->openpyxl只支持这种格式
from openpyxl import load_workbook  #专门用来打开excel的

#1.打开excel
wb=load_workbook("test.xlsx") #打开工作簿

#2.定位表单名，接下来要操作的表单名
sheet=wb['python'] #传表单名 返回一个表单对象

#3.定位单元格 行列值
res=sheet.cell(1,1).value

print("最大行:{}".format(sheet.max_row))   #求表单的最大行
print("最大列：{}".format(sheet.max_column)) #求表单的最大列

#print("拿到的结果是",res)

#数据从excel里面拿出来是什么类型
#数字还是数字，其他都是字符串
print("url:{},类型是{}").format(sheet.cell(1,1).value.type(sheet.cell(1,1).vaule))
print("url:{},类型是{}").format(sheet.cell(1,2).value.type(sheet.cell(1,2).vaule))
print("url:{},类型是{}").format(sheet.cell(1,3).value.type(sheet.cell(1,3).vaule))
print("url:{},类型是{}").format(sheet.cell(1,4).value.type(sheet.cell(1,4).vaule))



